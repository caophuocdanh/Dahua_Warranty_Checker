import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import requests
import json
import threading
import queue
import os
from datetime import datetime

# --- CONFIGURATION ---
API_URL = "https://app.dahua.vn:7778/Api.svc/Web/TraCuuBaoHanhTheoSeria?seria={serial}"

# --- Try to import openpyxl for Excel export ---
try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

class WarrantyCheckerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Dahua Warranty Checker")
        self.root.geometry("800x400")

        # Add icon
        try:
            icon_path = os.path.join(os.path.dirname(__file__), "icon.ico")
            if os.path.exists(icon_path):
                self.root.iconbitmap(icon_path)
            else:
                print(f"Warning: icon.ico not found at {icon_path}")
        except Exception as e:
            print(f"Error setting icon: {e}")

        self.filepath = ""
        self.result_queue = queue.Queue()

        # --- UI Elements ---
        self.create_widgets()
        self.process_queue()

    def create_widgets(self):
        # --- Top Frame for all controls and status ---
        top_frame = ttk.Frame(self.root, padding="10")
        top_frame.pack(fill='x')

        # Right-aligned controls are packed first
        self.export_button = ttk.Button(top_frame, text="Xuất Excel", command=self.export_to_excel)
        self.export_button.pack(side='right', padx=(10, 0))

        # Left-aligned controls
        ttk.Button(top_frame, text="Chọn file", command=self.select_file).pack(side='left', padx=(0, 10))
        self.file_label = ttk.Label(top_frame, text="Chưa chọn file nào")
        self.file_label.pack(side='left', padx=(0, 20))
        self.check_button = ttk.Button(top_frame, text="KIỂM TRA BẢO HÀNH", command=self.start_checking_thread)
        self.check_button.pack(side='left', padx=(0, 10))
        self.status_label = ttk.Label(top_frame, text="Trạng thái: Sẵn sàng")
        self.status_label.pack(side='left', fill='x', expand=True)

        # --- Treeview for results ---
        tree_frame = ttk.Frame(self.root, padding="10")
        tree_frame.pack(fill='both', expand=True)

        columns = ("stt", "sn", "ten", "so_thang_bh", "ngay_xuat", "ngay_con_lai", "so_lan_bh")
        self.tree = ttk.Treeview(tree_frame, columns=columns, show='headings')

        # Define headings
        self.tree.heading("stt", text="STT")
        self.tree.heading("sn", text="Serial Number")
        self.tree.heading("ten", text="Mã Sản Phẩm")
        self.tree.heading("so_thang_bh", text="Tháng BH")
        self.tree.heading("ngay_xuat", text="Ngày Xuất")
        self.tree.heading("ngay_con_lai", text="Ngày BH Còn Lại")
        self.tree.heading("so_lan_bh", text="Số Lần BH")

        # Define column widths
        self.tree.column("stt", width=40, anchor='center')
        self.tree.column("sn", width=120, stretch=False)
        self.tree.column("ten", width=180, stretch=False)
        self.tree.column("so_thang_bh", width=80, anchor='center')
        self.tree.column("ngay_xuat", width=120, anchor='center')
        self.tree.column("ngay_con_lai", width=120, anchor='center')
        self.tree.column("so_lan_bh", width=80, anchor='center')

        # Add scrollbars
        vsb = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        vsb.pack(side='right', fill='y')
        hsb.pack(side='bottom', fill='x')
        self.tree.pack(fill='both', expand=True)

        # --- Footer Frame for data source info ---
        footer_frame = ttk.Frame(self.root, padding="5")
        footer_frame.pack(fill='x')
        ttk.Label(footer_frame, text="* Dữ liệu bảo hành từ nhà phân phối DSS (dahua.vn).", foreground="gray").pack(side='left')

    def export_to_excel(self):
        if not OPENPYXL_AVAILABLE:
            messagebox.showerror("Thiếu thư viện", 
                                 "Chức năng xuất Excel yêu cầu thư viện 'openpyxl'.\n\n" 
                                 "Vui lòng cài đặt bằng cách mở Command Prompt hoặc Terminal và gõ lệnh:\n" 
                                 "pip install openpyxl")
            return

        if not self.tree.get_children():
            messagebox.showwarning("Không có dữ liệu", "Không có dữ liệu trong bảng để xuất.")
            return

        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        default_filename = f"Dahua_Warranty_Checker_{timestamp}.xlsx"

        filepath = filedialog.asksaveasfilename(
            initialfile=default_filename,
            defaultextension=".xlsx",
            filetypes=[("Excel Workbook", "*.xlsx"), ("All Files", "*.*")],
            title="Lưu file Excel"
        )
        if not filepath:
            return

        try:
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Warranty_Data"

            # Write header
            header = [self.tree.heading(col)['text'] for col in self.tree['columns']]
            sheet.append(header)

            # Define border style
            thin_border = Border(left=Side(style='thin'),
                                 right=Side(style='thin'),
                                 top=Side(style='thin'),
                                 bottom=Side(style='thin'))

            # Apply header formatting
            header_font = Font(bold=True, color="FFFFFF") # White text
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid") # Blue
            header_alignment = Alignment(horizontal="center", vertical="center")

            for col_idx, cell in enumerate(sheet[1], 1): # Iterate through cells in the first row
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border # Apply border to header

            # Write data rows
            for item_id in self.tree.get_children():
                row_values = self.tree.item(item_id)['values']
                sheet.append(row_values)
            
            # Auto-fit columns and set alignment/border for data
            for col_idx, column_data in enumerate(zip(*sheet.iter_rows(min_row=1, values_only=True)), 1):
                max_length = 0
                for row_idx in range(1, sheet.max_row + 1): # Iterate through all rows for max_length
                    cell_value = sheet.cell(row=row_idx, column=col_idx).value
                    try:
                        if cell_value is not None:
                            max_length = max(max_length, len(str(cell_value)))
                    except TypeError: # Handle non-string types
                        max_length = max(max_length, len(str(cell_value)))
                
                adjusted_width = (max_length + 2) # Add a small buffer
                sheet.column_dimensions[get_column_letter(col_idx)].width = adjusted_width

                for row_idx in range(2, sheet.max_row + 1): # Start from second row (data)
                    cell = sheet.cell(row=row_idx, column=col_idx)
                    if col_idx == 1: # STT column
                        cell.alignment = Alignment(horizontal="center")
                    else:
                        cell.alignment = Alignment(horizontal="left")
                    cell.border = thin_border # Apply border to data cells

            # Freeze header row
            sheet.freeze_panes = 'A2'

            workbook.save(filepath)
            messagebox.showinfo("Thành công", f"Đã xuất dữ liệu thành công đến file:\n{filepath}")

        except Exception as e:
            messagebox.showerror("Lỗi", f"Đã xảy ra lỗi khi xuất file Excel: {e}")

    def select_file(self):
        self.filepath = filedialog.askopenfilename(
            title="Chọn file serials",
            filetypes=(("Text files", "*.txt"), ("All files", "*.*"))
        )
        if self.filepath:
            self.file_label.config(text=os.path.basename(self.filepath))
            self.status_label.config(text=f"Đã chọn file: {os.path.basename(self.filepath)}")
        else:
            self.file_label.config(text="Chưa chọn file nào")

    def start_checking_thread(self):
        if not self.filepath:
            messagebox.showwarning("Chưa chọn file", "Vui lòng chọn một file .txt chứa danh sách serial.")
            return

        # Clear previous results
        for i in self.tree.get_children():
            self.tree.delete(i)

        self.check_button.config(state='disabled')
        self.status_label.config(text="Đang xử lý...")

        # Start the background thread
        threading.Thread(target=self.worker_task, daemon=True).start()

    def worker_task(self):
        """This function runs in a background thread to avoid freezing the UI."""
        try:
            with open(self.filepath, 'r') as f:
                serials = [line.strip() for line in f if line.strip()]
        except Exception as e:
            self.result_queue.put({'error': f"Lỗi đọc file: {e}"})
            self.result_queue.put('DONE')
            return

        if not serials:
            self.result_queue.put({'error': "File serial trống."})
            self.result_queue.put('DONE')
            return
            
        for index, serial in enumerate(serials):
            info = self.fetch_warranty_info(serial)
            info['stt'] = index + 1
            info['original_serial'] = serial # Pass original serial to the queue
            self.result_queue.put(info)
        
        self.result_queue.put('DONE')

    def fetch_warranty_info(self, serial):
        """Fetches warranty info for a single serial."""
        try:
            response = requests.get(API_URL.format(serial=serial), timeout=15)
            response.raise_for_status()
            
            # API returns a JSON object where the 'd' key contains a JSON string.
            # This requires two parsing steps.
            outer_dict = response.json()
            inner_json_string = outer_dict.get('d', '[]') # Use .get for safety
            data = json.loads(inner_json_string)

            if data and isinstance(data, list):
                return data[0]
            return {'SN': serial, 'Ten': 'Không tìm thấy thông tin'}
        except requests.exceptions.RequestException:
            return {'SN': serial, 'Ten': 'Lỗi mạng'}
        except json.JSONDecodeError:
            return {'SN': serial, 'Ten': 'Lỗi phản hồi từ server'}
        except Exception:
            return {'SN': serial, 'Ten': 'Lỗi không xác định'}

    def process_queue(self):
        """Checks the queue for results from the worker thread and updates the UI."""
        try:
            while True:
                result = self.result_queue.get_nowait()
                if result == 'DONE':
                    self.check_button.config(state='normal')
                    self.status_label.config(text=f"Hoàn thành! Đã kiểm tra {len(self.tree.get_children())} serial.")
                    return
                
                if 'error' in result:
                    messagebox.showerror("Lỗi", result['error'])
                    continue

                # Get and format the date
                ngay_xuat = result.get('NgayXuat', 'N/A')
                if isinstance(ngay_xuat, str) and 'T' in ngay_xuat:
                    ngay_xuat = ngay_xuat.split('T')[0]

                # Insert data into the treeview and get the new item's ID
                item_id = self.tree.insert("", 'end', values=(
                    result.get('stt', ''),
                    result.get('SoSeria') or result.get('original_serial', 'N/A'),
                    result.get('MaHangHoa', 'N/A'),
                    result.get('SoThangBaoHanh', 'N/A'),
                    ngay_xuat,
                    result.get('SoNgayBaoHanhConLai', 'N/A'),
                    result.get('SoLanBaoHanh', 'N/A')
                ))
                # Auto-scroll to the newly added item
                self.tree.see(item_id)

        except queue.Empty:
            pass # No new messages.
        finally:
            # Schedule to check the queue again in 100ms
            self.root.after(100, self.process_queue)


if __name__ == "__main__":
    root = tk.Tk()
    app = WarrantyCheckerApp(root)
    root.mainloop()